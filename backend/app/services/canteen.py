import uuid
from datetime import datetime, timezone
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
import openpyxl
import io

from app.models.canteen import CanteenRecord
from app.models.community import CommunityElder
from app.models.user import User
from app.models.community_event import CommunityEvent
from app.utils.llm_parser import parse_canteen_text


async def submit_canteen_record(
    db: AsyncSession,
    community_id: uuid.UUID,
    worker_id: uuid.UUID,
    raw_text: str | None = None,
    excel_bytes: bytes | None = None,
    source_format: str = "text",
) -> CanteenRecord:
    """提交食堂数据并触发 LLM 解析"""
    if excel_bytes:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(values_only=True):
                lines.append("\t".join(str(c) for c in row if c is not None))
        raw_text = "\n".join(lines)
        source_format = "excel"

    record = CanteenRecord(
        community_id=community_id,
        raw_text=raw_text or "",
        source_format=source_format,
        parse_status="pending",
        recorded_by=worker_id,
    )
    db.add(record)
    await db.commit()
    await db.refresh(record)

    stmt = (
        select(CommunityElder, User.nickname)
        .join(User, CommunityElder.elder_id == User.id)
        .where(CommunityElder.community_id == community_id)
    )
    result = await db.execute(stmt)
    elder_rows = result.all()
    elder_list = [
        {"id": str(e.elder_id), "name": name, "care_level": e.care_level}
        for e, name in elder_rows
    ]

    parsed = await parse_canteen_text(raw_text or "", elder_list)

    if parsed.get("fallback"):
        record.parse_status = "failed"
        await db.commit()
        await db.refresh(record)
        return record

    record.parsed_data = parsed
    record.parsed_at = datetime.now(timezone.utc)
    record.parse_status = "success"
    await db.commit()
    await db.refresh(record)

    await _generate_absent_events(db, community_id, parsed, elder_list)

    return record


async def _generate_absent_events(
    db: AsyncSession,
    community_id: uuid.UUID,
    parsed: dict,
    elder_list: list[dict],
):
    """为未就餐的 A/B 级老人自动创建事件"""
    care_map = {e["id"]: e["care_level"] for e in elder_list}
    attendees = parsed.get("attendees", [])

    for att in attendees:
        elder_id_str = att.get("elder_id")
        if not elder_id_str or att.get("present") is not False:
            continue

        care_level = care_map.get(elder_id_str, "C")
        severity = "urgent" if care_level == "A" else "warning" if care_level == "B" else "info"

        try:
            elder_uuid = uuid.UUID(elder_id_str)
        except (ValueError, TypeError):
            continue

        event = CommunityEvent(
            community_id=community_id,
            elder_id=elder_uuid,
            event_type="absent",
            source="canteen",
            description=f"{att.get('elder_name', '未知')} 今日{parsed.get('meal_type', '')}未就餐",
            severity=severity,
        )
        db.add(event)

    await db.commit()
