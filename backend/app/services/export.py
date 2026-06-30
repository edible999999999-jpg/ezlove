import io
import uuid
from datetime import datetime, date, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import select, func, and_, cast, Date
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.community import CommunityElder, CommunityWorker
from app.models.community_event import CommunityEvent
from app.models.canteen import CanteenRecord
from app.models.view_event import ViewEvent
from app.models.user import User

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="F5EFE7", end_color="F5EFE7", fill_type="solid")
HEADER_BORDER = Border(bottom=Side(style="thin", color="C4BAB0"))


def _style_header(ws):
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center")


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


async def export_elder_list(db: AsyncSession, community_id: uuid.UUID) -> bytes:
    stmt = (
        select(CommunityElder, User.nickname)
        .join(User, CommunityElder.elder_id == User.id)
        .where(CommunityElder.community_id == community_id)
        .order_by(CommunityElder.care_level, User.nickname)
    )
    result = await db.execute(stmt)
    rows = result.all()

    worker_ids = {r[0].assigned_worker_id for r in rows if r[0].assigned_worker_id}
    worker_map = {}
    if worker_ids:
        w_result = await db.execute(
            select(CommunityWorker.id, CommunityWorker.name)
            .where(CommunityWorker.id.in_(worker_ids))
        )
        worker_map = {wid: wname for wid, wname in w_result.all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "老人名单"
    ws.append(["姓名", "护理等级", "地址", "风险评分", "风险等级", "紧急联系人", "联系电话", "健康备注", "负责网格员"])
    _style_header(ws)

    for elder, nickname in rows:
        ws.append([
            nickname or "—",
            elder.care_level,
            elder.address or "—",
            elder.risk_score,
            elder.risk_level or "—",
            elder.emergency_contact_name or "—",
            elder.emergency_contact_phone or "—",
            elder.health_notes or "—",
            worker_map.get(elder.assigned_worker_id, "—"),
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    return _to_bytes(wb)


async def export_events(
    db: AsyncSession,
    community_id: uuid.UUID,
    start_date: date | None = None,
    end_date: date | None = None,
) -> bytes:
    stmt = select(CommunityEvent).where(CommunityEvent.community_id == community_id)
    if start_date:
        stmt = stmt.where(CommunityEvent.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        stmt = stmt.where(CommunityEvent.created_at < datetime.combine(end_date + timedelta(days=1), datetime.min.time()))
    stmt = stmt.order_by(CommunityEvent.created_at.desc())
    result = await db.execute(stmt)
    events = result.scalars().all()

    elder_ids = {e.elder_id for e in events}
    name_map = {}
    if elder_ids:
        n_result = await db.execute(select(User.id, User.nickname).where(User.id.in_(elder_ids)))
        name_map = {uid: name for uid, name in n_result.all()}

    type_labels = {"fall": "跌倒", "absent": "缺勤", "emergency": "紧急", "visit": "探访", "other": "其他", "manual_confirm": "手动确认"}
    severity_labels = {"urgent": "紧急", "warning": "警告", "info": "信息"}
    source_labels = {"canteen": "食堂", "alert": "预警", "manual": "手动"}

    wb = Workbook()
    ws = wb.active
    ws.title = "事件记录"
    ws.append(["时间", "老人姓名", "事件类型", "来源", "严重程度", "描述", "状态", "处理备注", "处理时间"])
    _style_header(ws)

    for e in events:
        ws.append([
            e.created_at.strftime("%Y-%m-%d %H:%M"),
            name_map.get(e.elder_id, "—"),
            type_labels.get(e.event_type, e.event_type),
            source_labels.get(e.source, e.source),
            severity_labels.get(e.severity, e.severity),
            e.description or "—",
            "已处理" if e.is_resolved else "待处理",
            e.resolution_note or "—",
            e.resolved_at.strftime("%Y-%m-%d %H:%M") if e.resolved_at else "—",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    return _to_bytes(wb)


async def export_canteen_records(
    db: AsyncSession,
    community_id: uuid.UUID,
    start_date: date | None = None,
    end_date: date | None = None,
) -> bytes:
    stmt = (
        select(CanteenRecord)
        .where(CanteenRecord.community_id == community_id, CanteenRecord.parse_status == "success")
    )
    if start_date:
        stmt = stmt.where(CanteenRecord.created_at >= datetime.combine(start_date, datetime.min.time()))
    if end_date:
        stmt = stmt.where(CanteenRecord.created_at < datetime.combine(end_date + timedelta(days=1), datetime.min.time()))
    stmt = stmt.order_by(CanteenRecord.created_at.desc())
    result = await db.execute(stmt)
    records = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "食堂出勤"
    ws.append(["日期", "餐次", "姓名", "是否出席", "护理等级", "备注"])
    _style_header(ws)

    for r in records:
        parsed = r.parsed_data or {}
        meal_type = parsed.get("meal_type", "—")
        date_str = r.created_at.strftime("%Y-%m-%d")
        for attendee in parsed.get("attendees", []):
            ws.append([
                date_str,
                meal_type,
                attendee.get("name", "—"),
                "出席" if attendee.get("present") else "缺席",
                attendee.get("care_level", "—"),
                attendee.get("note", "—"),
            ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    return _to_bytes(wb)


async def export_activity_summary(
    db: AsyncSession,
    community_id: uuid.UUID,
    days: int = 30,
) -> bytes:
    stmt = (
        select(CommunityElder.elder_id, User.nickname)
        .join(User, CommunityElder.elder_id == User.id)
        .where(CommunityElder.community_id == community_id)
        .order_by(CommunityElder.care_level, User.nickname)
    )
    result = await db.execute(stmt)
    elders = result.all()

    today = date.today()
    date_range = [today - timedelta(days=i) for i in range(days - 1, -1, -1)]

    active_data = {}
    for d in date_range:
        day_start = datetime.combine(d, datetime.min.time())
        day_end = datetime.combine(d + timedelta(days=1), datetime.min.time())
        active_stmt = (
            select(ViewEvent.viewer_id)
            .where(ViewEvent.viewed_at >= day_start, ViewEvent.viewed_at < day_end)
            .distinct()
        )
        active_result = await db.execute(active_stmt)
        active_ids = set(active_result.scalars().all())
        active_data[d] = active_ids

    wb = Workbook()
    ws = wb.active
    ws.title = "活跃度汇总"
    headers = ["姓名"] + [d.strftime("%m/%d") for d in date_range]
    ws.append(headers)
    _style_header(ws)

    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    for elder_id, nickname in elders:
        row = [nickname or "—"]
        for d in date_range:
            active = elder_id in active_data.get(d, set())
            row.append("✓" if active else "✗")
        ws.append(row)
        row_num = ws.max_row
        for col_idx, d in enumerate(date_range, start=2):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = green_fill if cell.value == "✓" else red_fill
            cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    for i in range(2, len(date_range) + 2):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 6

    return _to_bytes(wb)
